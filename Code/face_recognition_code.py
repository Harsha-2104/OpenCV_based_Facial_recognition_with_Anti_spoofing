import cv2
import numpy as np
import face_recognition
import os
from datetime import date
import time
import dlib
from scipy.spatial import distance
import openpyxl
from openpyxl import Workbook, load_workbook

# Flags & Counters for Liveness Checks
consecutive_successful_checks = 0
required_successful_checks = 1
real_face_hits = 0

# Load facial landmarks predictor
predictor = dlib.shape_predictor('shape_predictor_68_face_landmarks.dat')
detector = dlib.get_frontal_face_detector()

# Eye Aspect Ratio Calculation
def eye_aspect_ratio(eye):
    A = distance.euclidean(eye[1], eye[5])
    B = distance.euclidean(eye[2], eye[4])
    C = distance.euclidean(eye[0], eye[3])
    ear = (A + B) / (2.0 * C)
    return ear

EYE_AR_THRESHOLD = 0.215
LIP_DISTANCE_THRESHOLD = 10
HEAD_MOVEMENT_THRESHOLD = 10
last_nose_position = None

# Liveness check functions
def detect_eye_blink(landmarks):
    left_eye = np.array([(landmarks.part(i).x * 4, landmarks.part(i).y * 4) for i in range(36, 42)])
    right_eye = np.array([(landmarks.part(i).x * 4, landmarks.part(i).y * 4) for i in range(42, 48)])

    left_ear = eye_aspect_ratio(left_eye)
    right_ear = eye_aspect_ratio(right_eye)
    ear = (left_ear + right_ear) / 2.0

    print(f"EAR: {ear:.3f}")

    if ear < EYE_AR_THRESHOLD:
        return True
    return False

def detect_lip_movement(landmarks):
    top_lip = np.array([(landmarks.part(i).x * 4, landmarks.part(i).y * 4) for i in range(50, 53)] + [(landmarks.part(i).x * 4, landmarks.part(i).y * 4) for i in range(61, 64)])
    bottom_lip = np.array([(landmarks.part(i).x * 4, landmarks.part(i).y * 4) for i in range(56, 59)] + [(landmarks.part(i).x * 4, landmarks.part(i).y * 4) for i in range(65, 68)])

    top_mean = np.mean(top_lip, axis=0)
    bottom_mean = np.mean(bottom_lip, axis=0)

    distance_between_lips = abs(top_mean[1] - bottom_mean[1])

    return distance_between_lips > LIP_DISTANCE_THRESHOLD

def detect_head_movement(landmarks):
    global last_nose_position

    nose_point = (landmarks.part(30).x * 4, landmarks.part(30).y * 4)

    if last_nose_position is None:
        last_nose_position = nose_point
        return False

    distance_moved = distance.euclidean(nose_point, last_nose_position)
    last_nose_position = nose_point

    return distance_moved > HEAD_MOVEMENT_THRESHOLD

def run_liveness_checks(landmarks):
    # This requires all checks to be true for liveness
    return detect_eye_blink(landmarks) and detect_lip_movement(landmarks) and detect_head_movement(landmarks)

# Load known faces
CurrentFolder = os.getcwd()
images_folder = os.path.join(CurrentFolder, 'faces')
known_face_encodings = []
known_face_names = []

for file in os.listdir(images_folder):
    if file.endswith((".png", ".jpg", ".jpeg")):
        image_path = os.path.join(images_folder, file)
        image = face_recognition.load_image_file(image_path)
        face_encodings = face_recognition.face_encodings(image)

        if face_encodings:
            known_face_encodings.append(face_encodings[0])
            known_face_names.append(os.path.splitext(file)[0])

# Initialize Excel file using openpyxl
excel_file_name = 'attendence_excel.xlsx' # Changed to .xlsx

# Check if file exists, if not, create a new workbook
if not os.path.exists(excel_file_name):
    wb = Workbook()
    ws = wb.active # Get the active sheet
    ws.title = 'Sheet1' # Default sheet title
    wb.save(excel_file_name)

# Load existing workbook and get input for new sheet name
wb = load_workbook(excel_file_name)
inp = input('Please give file name: ')
sheet_name = str(date.today()) + ',' + inp

# Check if the sheet already exists, if not create it
if sheet_name not in wb.sheetnames:
    sheet1 = wb.create_sheet(sheet_name)
else:
    sheet1 = wb[sheet_name]

# Write headers if the sheet is new or empty
if sheet1.cell(row=1, column=1).value is None: # Check if the first cell is empty
    sheet1['A1'] = 'ID NO.'
    sheet1['B1'] = 'Attended Time'

# Find the next available row
row = sheet1.max_row + 1
if sheet1.cell(row=1, column=1).value is None: # If the sheet was just created and headers written, start from row 2
    row = 2

already_attendence_taken = []
# Load existing attendance data from the current sheet to prevent re-recording
for r in range(2, sheet1.max_row + 1): # Start from row 2 to skip headers
    name_in_sheet = sheet1.cell(row=r, column=1).value
    if name_in_sheet:
        already_attendence_taken.append(name_in_sheet)

# Video capture setup
video_capture = cv2.VideoCapture(0)
FRAME_PROCESSING_FREQUENCY = 5
DELAY_BETWEEN_FRAMES = 0.1
frame_counter = 0

while True:
    ret, frame = video_capture.read()
    if not ret:
        print("Failed to grab frame")
        break

    time.sleep(DELAY_BETWEEN_FRAMES)
    frame_counter += 1

    if frame_counter % FRAME_PROCESSING_FREQUENCY != 0:
        continue

    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
    gray = cv2.cvtColor(small_frame, cv2.COLOR_BGR2GRAY)

    face_locations = face_recognition.face_locations(rgb_small_frame)

    for (top, right, bottom, left) in face_locations:
        # Scale back the coordinates for dlib's detector as it works on original size
        # and then for drawing on the full-size frame
        top_orig, right_orig, bottom_orig, left_orig = top * 4, right * 4, bottom * 4, left * 4

        face_rect = dlib.rectangle(left, top, right, bottom) # Use scaled down coordinates for landmarks
        landmarks = predictor(gray, face_rect)

        # Draw rectangle around face (on the full-size frame)
        cv2.rectangle(frame, (left_orig, top_orig), (right_orig, bottom_orig), (0, 255, 0), 2)


        if run_liveness_checks(landmarks):
            # Only encode and compare if liveness check passes
            # Pass only the detected face location to face_encodings for efficiency
            face_encoding = face_recognition.face_encodings(rgb_small_frame, [(top, right, bottom, left)])[0]
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            if True in matches:
                best_match_index = np.argmin(face_recognition.face_distance(known_face_encodings, face_encoding))
                name = known_face_names[best_match_index]

            if name not in already_attendence_taken and name != "Unknown":
                sheet1.cell(row=row, column=1, value=name)
                sheet1.cell(row=row, column=2, value=time.ctime())
                row += 1
                already_attendence_taken.append(name)
                wb.save(excel_file_name)


            cv2.putText(frame, f"Real Face Detected: {name}", (left_orig, top_orig - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)
            print(f"Real Face Detected: {name}")
        else:
            cv2.putText(frame, "Fake Face Detected", (left_orig, top_orig - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 255), 2)

    cv2.imshow('Face Attendance System', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

wb.save(excel_file_name) # Save the workbook one last time
video_capture.release()
cv2.destroyAllWindows()
